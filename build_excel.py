"""
ClearMetric Freelance Rate Calculator — Premium Excel Template
Product for Gumroad ($11.99)

3 Sheets:
  1. Rate Calculator — inputs (gold), calculated rates, revenue breakdown
  2. Project Pricer — input project details, get quoted price with line items
  3. How To Use — instructions

Design: Gold/Amber palette (#B7950B primary, #7D6608 dark, #FEF9E7 input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
import os

# ============================================================
# DESIGN SYSTEM — Gold/Amber
# ============================================================
GOLD = "B7950B"
DARK_GOLD = "7D6608"
WHITE = "FFFFFF"
INPUT_AMBER = "FEF9E7"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"
RED = "E74C3C"
LIGHT_RED = "FDEDEC"
ACCENT = "D4AC0D"
LIGHT_GOLD = "FEF5E7"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="F9E79F", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=DARK_GOLD, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=DARK_GOLD)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)
FONT_CTA = Font(name="Calibri", size=12, bold=True, color=DARK_GOLD)

FILL_GOLD = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK_GOLD, end_color=DARK_GOLD, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_AMBER, end_color=INPUT_AMBER, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY), right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY), bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_GOLD
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_GOLD
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None, hint=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt
    if hint:
        ch = ws.cell(row=row, column=vc + 1, value=hint)
        ch.font = FONT_SMALL
        ch.alignment = ALIGN_L


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: RATE CALCULATOR
# ============================================================
def build_rate_calculator(ws):
    ws.title = "Rate Calculator"
    ws.sheet_properties.tabColor = GOLD
    cols(ws, {"A": 2, "B": 36, "C": 18, "D": 4, "E": 36, "F": 18, "G": 2})

    for r in range(1, 70):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:F1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:F2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="FREELANCE RATE CALCULATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="Enter your numbers in the gold cells. Rates update automatically.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== LEFT: INPUTS =====
    header_bar(ws, 5, 2, 3, "INCOME & EXPENSES")
    label_input(ws, 6, 2, 3, "Target Annual Income ($)", 80000, "$#,##0")
    label_input(ws, 7, 2, 3, "Annual Business Expenses ($)", 5000, "$#,##0")
    label_input(ws, 8, 2, 3, "Self-Employment Tax Rate", 0.153, "0.0%")
    label_input(ws, 9, 2, 3, "Effective Income Tax Rate", 0.22, "0.0%")

    header_bar(ws, 11, 2, 3, "BENEFITS & SAVINGS")
    label_input(ws, 12, 2, 3, "Health Insurance ($/month)", 500, "$#,##0")
    label_input(ws, 13, 2, 3, "Retirement Savings %", 0.10, "0%")

    header_bar(ws, 15, 2, 3, "TIME OFF")
    label_input(ws, 16, 2, 3, "Vacation Weeks/Year", 4, "0")
    label_input(ws, 17, 2, 3, "Sick/Personal Days/Year", 10, "0")
    label_input(ws, 18, 2, 3, "Holidays/Year", 10, "0")

    header_bar(ws, 20, 2, 3, "BILLABLE HOURS")
    label_input(ws, 21, 2, 3, "Billable Hours/Day", 6, "0.0")
    label_input(ws, 22, 2, 3, "Days/Week", 5, "0")

    header_bar(ws, 24, 2, 3, "MARGIN")
    label_input(ws, 25, 2, 3, "Desired Profit Margin %", 0.20, "0%")

    # ===== RIGHT: RESULTS =====
    header_bar(ws, 5, 5, 6, "RESULTS", FILL_DARK)

    # Billable time
    label_calc(ws, 6, 5, 6, "Billable Weeks/Year",
               "=52-(C16+(C17+C18)/5)", "0.0")
    label_calc(ws, 7, 5, 6, "Billable Hours/Year",
               "=F6*C22*C21", "$#,##0")
    label_calc(ws, 8, 5, 6, "Total Annual Costs",
               "=C6*C8+C6*C9+C12*12+C6*C13+C7", "$#,##0", bold=True)
    label_calc(ws, 9, 5, 6, "Revenue Needed (no margin)",
               "=C6+F8", "$#,##0")
    label_calc(ws, 10, 5, 6, "Revenue (with margin)",
               "=F9/(1-C25)", "$#,##0", bold=True)

    header_bar(ws, 12, 5, 6, "RATES")
    label_calc(ws, 13, 5, 6, "Minimum Hourly",
               "=F9/F7", "$#,##0")
    label_calc(ws, 14, 5, 6, "Recommended Hourly",
               "=F10/F7", "$#,##0", bold=True)
    label_calc(ws, 15, 5, 6, "Day Rate",
               "=F14*C21", "$#,##0")
    label_calc(ws, 16, 5, 6, "Monthly Retainer",
               "=F14*F7/12", "$#,##0")

    # Revenue breakdown
    header_bar(ws, 18, 5, 6, "REVENUE BREAKDOWN")
    label_calc(ws, 19, 5, 6, "Take-Home", "=C6", "$#,##0")
    label_calc(ws, 20, 5, 6, "Taxes", "=C6*C8+C6*C9", "$#,##0")
    label_calc(ws, 21, 5, 6, "Health Insurance", "=C12*12", "$#,##0")
    label_calc(ws, 22, 5, 6, "Retirement", "=C6*C13", "$#,##0")
    label_calc(ws, 23, 5, 6, "Business Expenses", "=C7", "$#,##0")
    label_calc(ws, 24, 5, 6, "Profit Margin", "=F10-F9", "$#,##0")

    # Pie chart data (rows 26-31 for breakdown)
    for r in range(26, 32):
        ws.cell(row=r, column=5, value="").fill = FILL_WHITE

    # Chart
    chart = PieChart()
    chart.title = "Where Your Revenue Goes"
    chart.style = 10
    data = Reference(ws, min_col=6, min_row=18, max_row=24)
    cats = Reference(ws, min_col=5, min_row=19, max_row=24)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 10
    ws.add_chart(chart, "B34")

    # Protection
    ws.protection.sheet = True
    input_cells = [(6, 3), (7, 3), (8, 3), (9, 3), (12, 3), (13, 3),
                   (16, 3), (17, 3), (18, 3), (21, 3), (22, 3), (25, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: PROJECT PRICER
# ============================================================
def build_project_pricer(wb):
    ws = wb.create_sheet("Project Pricer")
    ws.sheet_properties.tabColor = ACCENT
    rc = "'Rate Calculator'"
    cols(ws, {"A": 2, "B": 36, "C": 18, "D": 4, "E": 18, "F": 2})

    for r in range(1, 50):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    for r in range(1, 4):
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="PROJECT PRICER").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:E3")
    ws.cell(row=3, column=2, value="Quote projects with line items. Uses your recommended hourly rate.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    header_bar(ws, 5, 2, 4, "PROJECT INPUTS")
    label_input(ws, 6, 2, 3, "Project Name", "Website Redesign")
    label_input(ws, 7, 2, 3, "Estimated Hours", 40, "0")
    label_input(ws, 8, 2, 3, "Hourly Rate (or blank = recommended)", None, "$#,##0")
    ws.cell(row=8, column=3).value = f"={rc}!F14"
    ws.cell(row=8, column=3).font = FONT_INPUT
    ws.cell(row=8, column=3).fill = FILL_INPUT
    ws.cell(row=8, column=3).number_format = "$#,##0"
    ws.cell(row=8, column=3).border = THIN
    ws.cell(row=8, column=3).alignment = ALIGN_R

    header_bar(ws, 10, 2, 4, "LINE ITEMS")
    line_headers = ["Description", "Hours", "Rate", "Amount"]
    for i, h in enumerate(line_headers):
        cell = ws.cell(row=11, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_GOLD
        cell.alignment = ALIGN_C
        cell.border = THIN

    # Line items (5 rows)
    for i in range(5):
        r = 12 + i
        ws.cell(row=r, column=2, value=f"Phase {i+1}" if i < 3 else "").font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=3, value=10 if i < 3 else 0).font = FONT_INPUT
        ws.cell(row=r, column=3).fill = FILL_INPUT
        ws.cell(row=r, column=3).number_format = "0"
        ws.cell(row=r, column=3).border = THIN
        ws.cell(row=r, column=4, value=f"={rc}!F14").font = FONT_VALUE
        ws.cell(row=r, column=4).fill = FILL_WHITE
        ws.cell(row=r, column=4).number_format = "$#,##0"
        ws.cell(row=r, column=4).border = THIN
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}").font = FONT_BOLD
        ws.cell(row=r, column=5).fill = FILL_WHITE
        ws.cell(row=r, column=5).number_format = "$#,##0"
        ws.cell(row=r, column=5).border = THIN

    header_bar(ws, 18, 2, 4, "Total Project Price")
    ws.cell(row=18, column=5, value="=SUM(E12:E16)").font = FONT_CTA
    ws.cell(row=18, column=5).number_format = "$#,##0"
    ws.cell(row=18, column=5).alignment = ALIGN_R
    ws.cell(row=18, column=5).fill = FILL_LIGHT
    ws.cell(row=18, column=5).border = THIN

    ws.protection.sheet = True
    for r in range(12, 17):
        ws.cell(row=r, column=2).protection = openpyxl.styles.Protection(locked=False)
        ws.cell(row=r, column=3).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE FREELANCE RATE CALCULATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Rate Calculator' tab and enter your numbers in the GOLD cells",
            "2. Results appear on the right: hourly rate, day rate, monthly retainer",
            "3. Check the revenue breakdown and pie chart",
            "4. Use 'Project Pricer' to quote projects with line items",
        ]),
        ("INPUT EXPLANATIONS", [
            "Target Annual Income: What you want to take home after taxes and expenses",
            "Annual Business Expenses: Software, tools, coworking, marketing, etc.",
            "Self-Employment Tax: Typically 15.3% in the US (SS + Medicare)",
            "Effective Income Tax: Your marginal/federal+state rate",
            "Health Insurance: Monthly premium (you pay as freelancer)",
            "Retirement Savings: % of income to save (e.g., 10%)",
            "Vacation/Sick/Holidays: Days off reduce billable time",
            "Billable Hours/Day: Not all 8 hours are billable — admin, meetings, etc.",
            "Days/Week: Usually 5; adjust if part-time",
            "Profit Margin: Buffer for slow months, raises, emergencies (e.g., 20%)",
        ]),
        ("INTERPRETING RESULTS", [
            "Minimum Hourly: Covers costs but no profit margin",
            "Recommended Hourly: Includes your profit margin — use this for quotes",
            "Day Rate: Recommended hourly × billable hours per day",
            "Monthly Retainer: Useful for ongoing clients",
            "Revenue Breakdown: See where your revenue goes (taxes, insurance, etc.)",
        ]),
        ("PROJECT PRICER", [
            "Enter project name and estimated hours",
            "Add line items by phase or task",
            "Hourly rate defaults to your recommended rate from the calculator",
            "Total updates automatically. Use for client proposals.",
        ]),
        ("IMPORTANT NOTES", [
            "Tax rates are estimates — consult a CPA for your situation",
            "Billable hours vary by industry; 5-6 hrs/day is typical for knowledge work",
            "Profit margin protects against slow months and client churn",
            "© 2026 ClearMetric. For educational use only. Not financial advice.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=DARK_GOLD)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Rate Calculator sheet...")
    build_rate_calculator(ws)

    print("Building Project Pricer sheet...")
    build_project_pricer(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "output", "ClearMetric-Freelance-Rate-Calculator.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
